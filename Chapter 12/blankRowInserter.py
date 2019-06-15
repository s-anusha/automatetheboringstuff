#! python3
# blankRowInserter.py
'''
Create a program blankRowInserter.py that takes two integers and a filename string as command line arguments. Let's call the first integer N and the second interger M. Starting at row N, the program should insert M blank rows into the spreadsheet.
You can write this program by reading in the contents of the spreadsheet. Then, when writing out the new spreadsheet, use a for loop to copy the first N lines, for the remaining lines, add M to the row number in the output spreadsheet.
'''
# Usage: python blankRowInserter.py rowNumber numberOfBlankRows fileName

import openpyxl
import sys

if len(sys.argv) < 4:
  print('Usage: python blankRowInserter.py rowNumber numberOfBlankRows fileName')
  sys.exit()

n = int(sys.argv[1])
m = int(sys.argv[2])
file = sys.argv[3]

if n < 1 or m < 1:
  print('Invalid input.')
  sys.exit()

wbRead = openpyxl.load_workbook(file)
readSheet = wbRead.active

wbWrite = openpyxl.Workbook()
writeSheet = wbWrite.active

for rowOfReadCellObjects in readSheet[readSheet.cell(row=1, column=1).coordinate:readSheet.cell(row=readSheet.max_row, column=readSheet.max_column).coordinate]:
  for readCell in rowOfReadCellObjects:
      if readCell.row < m - 1:
        writeSheet.cell(row=readCell.row, column=readCell.column).value = readCell.value
      else:
        writeSheet.cell(row=readCell.row+m, column=readCell.column).value = readCell.value

wbWrite.save('blankRowInserted.xlsx')

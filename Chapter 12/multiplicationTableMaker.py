#! python3
# multiplicationTableMaker.py
'''
Create a program multiplicationTable.py that takes a number N from the command line and creates an N×N multiplication table in an Excel spreadsheet.

Row 1 and column A should be used for labels and should be in bold.
'''
# Usage: python multiplicationTableMaker.py number

import openpyxl
from openpyxl.styles import Font
import sys

if len(sys.argv) < 2:
    print('Usage: python multiplicationTableMaker.py number')
    sys.exit()

n = int(sys.argv[1])

if n < 1:
    print('Invalid input.')
    sys.exit()

wb = openpyxl.Workbook()

activeSheet = wb.active

boldObj = Font(bold=True)

for i in range(1, n+1):
    activeSheet.cell(row=1, column=i+1).value = i
    activeSheet.cell(row=1, column=i+1).font = boldObj
    activeSheet.cell(row=i + 1, column=1).value = i
    activeSheet.cell(row=i + 1, column=1).font = boldObj

for i in range(1, n+1):
    for j in range(1, n+1):
        activeSheet.cell(row=i+1, column=j+1).value = activeSheet.cell(row=i+1, column=1).value * activeSheet.cell(row=1, column=j+1).value

wb.save('multiplicationTable.xlsx')

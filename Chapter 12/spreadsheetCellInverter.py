#! python3
# spreadsheetCellInverter.py
'''
Write a program to invert the row and column of the cells in the spreadsheet. For example, the value at row 5, column 3 will be at row 3, column 5 (and vice versa). This should be done for all cells in the spreadsheet.
You can write this program by using nested for loops to read in the spreadsheet's data into a list of lists data structure. This data structure could have sheetData[x][y] for the cell at column x and row y. Then, when writing out the new spreadsheet, use sheetData[y][x] for the cell at column x and row y.
'''
# Usage: python spreadsheetCellInverter.py fileName

import openpyxl
import sys

if len(sys.argv) < 2:
  print('Usage: python spreadsheetCellInverter.py fileName')
  sys.exit()

file = sys.argv[1]

wbRead = openpyxl.load_workbook(file)
readSheet = wbRead.active

wbWrite = openpyxl.Workbook()
writeSheet = wbWrite.active

for rowOfReadCellObjects in readSheet[readSheet.cell(row=1, column=1).coordinate:readSheet.cell(row=readSheet.max_row, column=readSheet.max_column).coordinate]:
  for readCell in rowOfReadCellObjects:
      writeSheet.cell(row=readCell.row, column=readCell.column).value = readSheet.cell(row=readCell.column, column=readCell.row).value

wbWrite.save('spreadsheetCellInverted.xlsx')

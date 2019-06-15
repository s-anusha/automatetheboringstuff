#! python3
# spreadsheetToTextFiles.py
'''
Write a program that performs the tasks of the previous program in reverse order: The program should open a spreadsheet and write the cells of column A into one text file, the cells of column B into another text file, and so on.
'''
# Usage: python spreadsheetToTextFiles.py fileName

import openpyxl
import sys

if len(sys.argv) < 2:
  print('Usage: python spreadsheetToTextFiles.py fileName')
  sys.exit()

file = sys.argv[1]

wbRead = openpyxl.load_workbook(file)
readSheet = wbRead.active
columnNumber = 1

for i in range(1, readSheet.max_column+1):
    file = open(str(columnNumber) + '.txt', 'w')
    for j in range(1, readSheet.max_row+1):
        if readSheet.cell(row=j, column=i).value is not None:
            file.write(str(readSheet.cell(row=j, column=i).value))
    file.close()
    columnNumber += 1
